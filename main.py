import os

from openpyxl import load_workbook
from openpyxl.cell import MergedCell


class Parser(object):
    def __init__(self, file_path):
        self._file_path = file_path
        self._merge_cells = {}
        self._html = ''
        self._data = []

    def run(self, title='打印页面'):
        wb = load_workbook(self._file_path, read_only=False)
        ws = wb.active
        self._merge_cells = {(item.min_row, item.min_col): (item.max_row, item.max_col) for item in
                             ws.merged_cells.ranges}
        self._html = '<html lang="zh">\n' \
                     '<head>\n' \
                     f'\t<title>{title}</title>\n' \
                     '\t<style>.topBtn{position: fixed;top: 5rem;right: 0.8rem;width: 3.2rem;height: 2.2rem;background-size: 100% auto;z-index: 9999;-webkit-transition:  opacity .3s ease;}</style>\n' \
                     '</head>\n' \
                     '<body>\n' \
                     '\t<table border="1" cellspacing="0" cellpadding="0" border-collapse="collapse">\n'
        for row in range(1, ws.max_row):
            self._html += '\t\t<tr>\n'
            for col in range(1, ws.max_column):
                cell = ws.cell(row, col)
                if not isinstance(cell, MergedCell):
                    cell_end = self._merge_cells.get((row, col))
                    if cell_end is None:
                        cell_end = (row, col)
                    self._html += f'\t\t\t<td rowspan="{cell_end[0] - row + 1}" colspan="{cell_end[1] - col + 1}">' \
                                  f'{cell.value}' \
                                  f'</td>\n'
                    self._data.append({
                        'value': cell.value,
                        'start': (row, col),
                        'end': cell_end
                    })
            self._html += '\t\t</tr>\n'
        self._html += '\t</table>\n' \
                      '<button class="topBtn">打印</button>\n' \
                      '</body>\n' \
                      '</html>'
        return self._html

    @property
    def data(self):
        return self._data

    @property
    def html(self):
        return self._html


def main():
    excel_dir = os.path.abspath('excels')
    if not (os.path.exists('htmls') and os.path.isdir('htmls')):
        os.mkdir('htmls')
    for file in os.listdir(excel_dir):
        parser = Parser(os.path.join(excel_dir, file))
        parser.run()
        with open(f'htmls/{file}.html', 'w', encoding='utf-8') as f:
            f.write(parser.html)


if __name__ == '__main__':
    main()
